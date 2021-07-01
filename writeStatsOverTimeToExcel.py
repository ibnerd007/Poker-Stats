import openpyxl
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.chart.label import DataLabelList

from search import *
import time

def writeStatsOverTimetoExcel(vpipM, pfrM, tbpM, cbpM, afM, afqM, wtsdM, wasdRelM, mwas, mwbs, 
							  playerIndices, dateFormat, handTypeDesired):
	# Keeps a running bankroll of regular players across multiple sessions
	# Players tracked:
	# Fish, Raymond, Scott, Cedric

	# Read:    playerName = sheet.cell(row=2, column=1).value
	# Write:   sheet.cell(row=i+2, column=1, value=playerDict[playerIDs[i]])

	""" PLEASE READ: sheet.max_row does not function as expected. Even if there is no value
	in an entire row, those cells can still trigger the max_row object. To reduce max_row, you
	must right click and delete the entire row(s), NOT just the content by pressing 'Delete'.
	Charts and graphs should move up correspondingly; this is how you know you have successfully
	lowered the max_row object. This can cause plenty of errors so be careful when you use it. """

	# 1. Open workbook and sheet -------------------------------------------------------------

	wb_path = r'Outputs\stats over time.xlsx'

	wb = openpyxl.load_workbook(wb_path) # load existing workbook

	if handTypeDesired == 'NL':
		sheet = wb['NL SOT'] # access sheet
	elif handTypeDesired == 'PLO':
		sheet = wb['PLO SOT'] # access sheet
	else: # combined hand types are desired
		sheet = wb['combined SOT'] # access sheet

	# 2. Check to make sure this session has not already been recorded -----------------------
	#    Compare this session's date with date column

	dates = []
	vpipCols = (3, 13, 23, 33) # columns in Excel where the net will be stored

	for i in range(2, sheet.max_row+1):
		cellDate = sheet.cell(row=i, column=1).value
		dates.append(cellDate)

	if search(dates, dateFormat) != -1: # data from this date has been entered previously
		return
	else: # add date to column
		print('Adding stats over time data...\n')
		if dates == [None]:
			sheet.cell(row=sheet.max_row, column=1, value=dateFormat)
		else:
			sheet.cell(row=sheet.max_row + 1, column=1, value=dateFormat)

	# 3. Add stats for each player -----------------------------------------------

	for i, pI in enumerate(playerIndices):

		playerStats = (vpipM[pI][2], pfrM[pI][2], tbpM[pI][2], afqM[pI][2], 
			wtsdM[pI][2], wasdRelM[pI][2], cbpM[pI][2], afM[pI][2]) # stats are across all positions

		if pI != -1: # player played the session

			for j, stat in enumerate(playerStats):
				# set each stat one by one in a nested loop
				if stat == -1: sheet.cell(row=sheet.max_row, column=vpipCols[i] + j, value=0)
				else:          sheet.cell(row=sheet.max_row, column=vpipCols[i] + j, value=stat)
				
				if j != 7: # NOT aggression factor
					sheet.cell(row=sheet.max_row, column=vpipCols[i] + j).number_format = '0.0%'
				else:
					sheet.cell(row=sheet.max_row, column=vpipCols[i] + j).number_format = '0.00'

		else: # player didn't play this session
			sheet.cell(row=sheet.max_row, column=vpipCols[i]-1, value='Didn\'t play')

	# Initialize bar chart variables ------------------------------------------------------------
	sheetname = ''
	
	if handTypeDesired == 'NL':
		chartsheet = wb['NL charts'] # access sheet
		sheetname = 'NL charts'

	elif handTypeDesired == 'PLO':
		chartsheet = wb['PLO charts'] # access sheet
		sheetname = 'PLO charts'

	else: # combined hand types are desired
		chartsheet = wb['combined charts'] # access sheet
		sheetname = 'combined charts'

	wb.remove(chartsheet)

	# After deleting old sheet, create new sheet with same name, put at penultimate index
	chartsheet = wb.create_sheet(sheetname, -1)

	first_column_idx = 'A'
	second_column_idx = 'T'
	width = 40
	height = 10
	style = 2
	shape = 4

	# Add bar charts -----------------------------------------------------------------------------
	
	chart1 = BarChart()
	chart1.type = "col"
	chart1.style = style
	chart1.title = "VPIP"

	chart1.width = width
	chart1.height = height
	chart1.shape = shape 

	chart1.legend.position = 'r'
	chart1.x_axis.title = 'Date'

	data_fish =    Reference(sheet, min_col=vpipCols[0], min_row=2, max_row=sheet.max_row)
	data_raymond = Reference(sheet, min_col=vpipCols[1], min_row=2, max_row=sheet.max_row)
	data_scott =   Reference(sheet, min_col=vpipCols[2], min_row=2, max_row=sheet.max_row)
	data_cedric =  Reference(sheet, min_col=vpipCols[3], min_row=2, max_row=sheet.max_row)

	series_fish =    Series(data_fish,    title='Fish')
	series_raymond = Series(data_raymond, title='Raymond')
	series_scott =   Series(data_scott,   title='Scott')
	series_cedric =  Series(data_cedric,  title='Cedric')

	chart1.append(series_fish)
	chart1.append(series_raymond)
	chart1.append(series_scott)
	chart1.append(series_cedric)

	cats = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
	chart1.set_categories(cats)

	chartsheet.add_chart(chart1, "{}2".format(first_column_idx))

	# ------------------------------------------------------------------------------------------

	chart1 = BarChart()
	chart1.type = "col"
	chart1.style = style
	chart1.title = "Aggression Frequency"

	chart1.width = width
	chart1.height = height
	chart1.shape = shape 

	chart1.legend.position = 'r'

	data_fish =    Reference(sheet, min_col=vpipCols[0] + 3, min_row=2, max_row=sheet.max_row)
	data_raymond = Reference(sheet, min_col=vpipCols[1] + 3, min_row=2, max_row=sheet.max_row)
	data_scott =   Reference(sheet, min_col=vpipCols[2] + 3, min_row=2, max_row=sheet.max_row)
	data_cedric =  Reference(sheet, min_col=vpipCols[3] + 3, min_row=2, max_row=sheet.max_row)

	series_fish =    Series(data_fish,    title='Fish')
	series_raymond = Series(data_raymond, title='Raymond')
	series_scott =   Series(data_scott,   title='Scott')
	series_cedric =  Series(data_cedric,  title='Cedric')

	chart1.append(series_fish)
	chart1.append(series_raymond)
	chart1.append(series_scott)
	chart1.append(series_cedric)

	cats = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
	chart1.set_categories(cats)

	chartsheet.add_chart(chart1, "{}23".format(first_column_idx))

	# ---------------------------------------------------------------------------------------------

	chart1 = BarChart()
	chart1.type = "col"
	chart1.style = style
	chart1.title = "Won at showdown"

	chart1.width = width
	chart1.height = height
	chart1.shape = shape 

	chart1.legend.position = 'r'

	data_fish =    Reference(sheet, min_col=vpipCols[0] + 5, min_row=2, max_row=sheet.max_row)
	data_raymond = Reference(sheet, min_col=vpipCols[1] + 5, min_row=2, max_row=sheet.max_row)
	data_scott =   Reference(sheet, min_col=vpipCols[2] + 5, min_row=2, max_row=sheet.max_row)
	data_cedric =  Reference(sheet, min_col=vpipCols[3] + 5, min_row=2, max_row=sheet.max_row)

	series_fish =    Series(data_fish,    title='Fish')
	series_raymond = Series(data_raymond, title='Raymond')
	series_scott =   Series(data_scott,   title='Scott')
	series_cedric =  Series(data_cedric,  title='Cedric')

	chart1.append(series_fish)
	chart1.append(series_raymond)
	chart1.append(series_scott)
	chart1.append(series_cedric)

	cats = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
	chart1.set_categories(cats)

	chartsheet.add_chart(chart1, "{}44".format(first_column_idx))

	# --------------------------------------------------------------------------------------------

	chart1 = BarChart()
	chart1.type = "col"
	chart1.style = style
	chart1.title = "C-bet %"

	chart1.width = width
	chart1.height = height
	chart1.shape = shape 

	chart1.legend.position = 'r'

	data_raymond = Reference(sheet, min_col=vpipCols[1] + 6, min_row=2, max_row=sheet.max_row)
	data_scott =   Reference(sheet, min_col=vpipCols[2] + 6, min_row=2, max_row=sheet.max_row)

	series_raymond = Series(data_raymond, title='Raymond')
	series_scott =   Series(data_scott,   title='Scott')

	chart1.append(series_raymond)
	chart1.append(series_scott)

	cats = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
	chart1.set_categories(cats)

	chartsheet.add_chart(chart1, "{}65".format(first_column_idx))

	# --------------------------------------------------------------------------------------------

	wb.save(wb_path)