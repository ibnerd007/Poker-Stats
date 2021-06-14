import openpyxl

wb_path = r'Outputs\bankrolls.xlsx'

wb = openpyxl.load_workbook(wb_path) # load existing workbook

sheet = wb['Bankrolls'] # access sheet

dates = []
for i in range(2, sheet.max_row + 1):
	cellDate = sheet.cell(row=i, column=1).value
	dates.append(cellDate)

print(dates)