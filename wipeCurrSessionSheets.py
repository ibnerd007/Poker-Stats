import openpyxl

def wipeCurrSessionSheets(handTypeDesired):
	""" Wipes sheets on stats.xlsx that are unused from current session.
	This provides clarity and avoids confusion, so only one date is presented at a time
	on stats.xlsx. """

	wb_path = r'Outputs\stats.xlsx'
	wb = openpyxl.load_workbook(wb_path) # load existing workbook

	sheetnames = wb.sheetnames

	sheetname = '{} Stats-this session'.format(handTypeDesired)

	if sheetname in sheetnames:
		sheet = wb[sheetname]
		wb.remove(sheet)

	wb.save(wb_path)