import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.chart.label import DataLabelList

def writeCurrSessionToExcel(vpipM, pfrM, tbpM, cbpCountM, afM, afqM, wtsdM, wasdM, mwas, mwbs, 
			                ledgerM, playerIDs, playerDict, handsPlayed, bestHandsM, date,
			                handTypeDesired):
	
	# wb = openpyxl.Workbook() # create new workbook
	wb_path = r'Outputs\stats.xlsx'
	wb = openpyxl.load_workbook(wb_path) # load existing workbook

	sheetname = '{} Stats-this session'.format(handTypeDesired)

	# Get list of sheets
	sheets = wb.sheetnames

	# Delete old sheet if it exists, to avoid stacking charts
	if sheetname in sheets:
		sheet = wb[sheetname] # Define sheet
		wb.remove(sheet)

	# Create new sheet with same name, put at first index
	sheet = wb.create_sheet(sheetname, 0)

	# sheet.insert_rows(1, 2) # Before 1st row, insert 2 columns

	# row = player
	# column = stat index
	# value = stat value

	column_headers = ('Player', 'Buy in', 'Buy out', 'Net', 'Rebuys', 'VPIP', 'Pre-flop Raise',
		'Three-bet', 'Aggro Frequency', 'Went to showdown', 'Won at showdown', 'Aggression Factor',
		'C-bets', 'C-bet opportunities', 'At showdown', 'Before showdown', 'Hands played')

	# Fill column headers on new sheet
	for (i, column_name) in enumerate(column_headers):
		sheet.cell(row=1, column=i+1, value=column_name)

	for i in range(len(playerIDs)): # Fill out player names first
		sheet.cell(row=i + 2, column=1, value=playerDict[playerIDs[i]])

	# Next, fill ledger stats
	for player in range(len(playerIDs)): # cols
		for stat in range(len(ledgerM[0])): # rows
			sheet.cell(row=player + 2, column=stat + 2, value=ledgerM[player][stat]) # averaged positions


	tdStats = [vpipM, pfrM, tbpM, afqM, wtsdM, wasdM]
	moneyStats = [mwas, mwbs]

	# Fill Excel spreadsheet with percent stat data
	for stat in range(len(tdStats)): # cols
		for player in range(len(playerIDs)): # rows
			sheet.cell(row=player + 2, column=stat + 6, value=tdStats[stat][player][2]) # averaged positions
			sheet.cell(row=player + 2, column=stat + 6).number_format = '0.0%'


	# Fill Excel spreadsheet with C-bets vs opportunities and aggression factors, not a percent-based stats
	for player in range(len(playerIDs)): # rows
		sheet.cell(row=player + 2, column=12, value=afM[player][2])
		
		totalBets = cbpCountM[player][0] + cbpCountM[player][1]
		sheet.cell(row=player + 2, column=13, value=totalBets) # fill c-bets

		totalOpps = cbpCountM[player][2] + cbpCountM[player][3]
		sheet.cell(row=player + 2, column=14, value=totalOpps) # fill c-bet opportunities


	# Fill Excel spreadsheet with monetary stat data
	for stat in range(len(moneyStats)):
		for player in range(len(playerIDs)):
			sheet.cell(row=player + 2, column=stat + 15, value=moneyStats[stat][player]) # averaged positions
			sheet.cell(row=player + 2, column=stat + 15).number_format = '"$"#,##0.00_-'

	for player in range(len(playerIDs)):
		totalHandsPlayed = handsPlayed[0][player] + handsPlayed[1][player]
		sheet.cell(row=player + 2, column=17, value=totalHandsPlayed) # averaged positions

	row = player + 2 + 1
	rows = sheet.max_row

	sheet.delete_rows(row, sheet.max_row) # delete rows that may remain from previous sessions

	sheet.cell(row=2, column=20, value='Date') # date of session
	sheet.cell(row=3, column=20, value=date) # date of session

	# Create 6 separate charts for data
	# Define chart formatting parameters

	width = 22
	height = round((9/16)*width, 1)
	style = 2 # standard
	shape = 4 # idk what this is
	first_column_idx = 'A' # Column 1
	second_column_idx = 'O' # Column 15

	# Pre-flop stats --------------------------------------------------------------------------------

	chart1 = BarChart()
	chart1.type = "col"
	chart1.style = style
	chart1.title = "VPIP, Pre-flop raise, 3-bet"

	chart1.width = width
	chart1.height = height

	chart1.legend.position = 'b'
	chart1.legend = None
	# chart1.x_axis.title = ''

	data = Reference(sheet, min_col=6, min_row=1, max_row=len(playerIDs)+1, max_col=8)
	cats = Reference(sheet, min_col=1, min_row=2, max_row=len(playerIDs)+1)
	chart1.add_data(data, titles_from_data=True)
	chart1.set_categories(cats)
	chart1.shape = shape 
	sheet.add_chart(chart1, "{}14".format(first_column_idx))	

	# Aggression factor -----------------------------------------------------------------------------

	chart1 = BarChart()
	chart1.type = "col"
	chart1.style = style
	chart1.title = "Aggression Factor"
	chart1.legend = None

	chart1.width = width
	chart1.height = height

	chart1.dataLabels = DataLabelList()
	chart1.dataLabels.showVal = True

	data = Reference(sheet, min_col=12, min_row=1, max_row=len(playerIDs)+1, max_col=12)
	cats = Reference(sheet, min_col=1, min_row=2, max_row=len(playerIDs)+1)
	chart1.add_data(data, titles_from_data=True)
	chart1.set_categories(cats)
	chart1.shape = shape
	sheet.add_chart(chart1, "{}14".format(second_column_idx))

	# C-bets vs opportunities -----------------------------------------------------------------------

	chart1 = BarChart()
	chart1.type = "col"
	chart1.style = style
	chart1.title = "C-bets vs opportunities"

	chart1.dataLabels = DataLabelList()
	chart1.dataLabels.showVal = True

	chart1.width = width
	chart1.height = height

	chart1.legend.position = 'b'

	data = Reference(sheet, min_col=13, min_row=1, max_row=len(playerIDs)+1, max_col=14)
	cats = Reference(sheet, min_col=1, min_row=2, max_row=len(playerIDs)+1)
	chart1.add_data(data, titles_from_data=True)
	chart1.set_categories(cats)
	chart1.shape = shape
	sheet.add_chart(chart1, "{}40".format(first_column_idx))

	# WTSD vs WASD ----------------------------------------------------------------------------------

	chart1 = BarChart()
	chart1.type = "col"
	chart1.style = style
	chart1.title = "Went to showdown vs Won at showdown"

	chart1.dataLabels = DataLabelList()
	chart1.dataLabels.showVal = True
	chart1.legend = None

	chart1.width = width
	chart1.height = height

	data = Reference(sheet, min_col=10, min_row=1, max_row=len(playerIDs)+1, max_col=11)
	cats = Reference(sheet, min_col=1, min_row=2, max_row=len(playerIDs)+1)
	chart1.add_data(data, titles_from_data=True)
	chart1.set_categories(cats)
	chart1.shape = shape
	sheet.add_chart(chart1, "{}40".format(second_column_idx))

	# MWAS vs MWBS ---------------------------------------------------------------------------------

	chart1 = BarChart()
	chart1.type = "col"
	chart1.style = style
	chart1.title = "Money won"

	chart1.dataLabels = DataLabelList()
	chart1.dataLabels.showVal = True
	chart1.legend.position = 'b'
	
	chart1.width = width
	chart1.height = height

	data = Reference(sheet, min_col=15, min_row=1, max_row=len(playerIDs)+1, max_col=16)
	cats = Reference(sheet, min_col=1, min_row=2, max_row=len(playerIDs)+1)
	chart1.add_data(data, titles_from_data=True)
	chart1.set_categories(cats)
	chart1.shape = shape
	sheet.add_chart(chart1, "{}66".format(first_column_idx))

	# # Hands played --------------------------------------------------------------------------------

	chart1 = BarChart()
	chart1.type = "bar"
	chart1.style = style
	chart1.title = "Hands played"

	chart1.dataLabels = DataLabelList()
	chart1.dataLabels.showVal = True
	chart1.legend = None

	chart1.width = 13
	chart1.height = 13

	data = Reference(sheet, min_col=17, min_row=1, max_row=len(playerIDs)+1, max_col=17)
	cats = Reference(sheet, min_col=1, min_row=2, max_row=len(playerIDs)+1)
	chart1.add_data(data, titles_from_data=True)
	chart1.set_categories(cats)
	chart1.shape = shape
	sheet.add_chart(chart1, "{}66".format(second_column_idx))

	# ---------------------------------------------------------------------------------------------

	wb.save(wb_path)




