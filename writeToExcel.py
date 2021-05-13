import openpyxl

def writeToExcel(vpipM, pfrM, tbpM, afM, afqM, wtsdM, wasdM, mwas, mwbs, 
			     ledgerM, staticIDs, playerIDs, playerDict, handsPlayed, bestHandsM):
	
	# wb = openpyxl.Workbook() # create new workbook
	wb_path = r'Outputs\stats.xlsx'

	wb = openpyxl.load_workbook(wb_path) # load existing workbook

	sheet = wb['Stats-this session'] # access sheet

	# for bankroll in bankrolls:
	# 	sheet.append(bankroll) # append to end of data

	# sheet.insert_rows(1, 2) # Before 1st row, insert 2 columns

	# row = player
	# column = stat index
	# value = stat value

	for i in range(len(playerIDs)): # Fill out player names first
		sheet.cell(row=i + 2, column=1, value=playerDict[playerIDs[i]])

	# Next, fill ledger stats
	for player in range(len(playerIDs)): # cols
		for stat in range(len(ledgerM[0])): # rows
			sheet.cell(row=player + 2, column=stat + 2, value=ledgerM[player][stat]) # averaged positions
			# print(ledgerM[player][stat])


	tdStats = [vpipM, pfrM, tbpM, afM, afqM, wtsdM, wasdM]
	moneyStats = [mwas, mwbs]

	# Fill Excel spreadsheet with percent stat data
	for stat in range(len(tdStats)): # rows
		for player in range(len(playerIDs)): # cols
			sheet.cell(row=player + 2, column=stat + 6, value=tdStats[stat][player][2]) # averaged positions

	# Fill Excel spreadsheet with monetary stat data
	for stat in range(len(moneyStats)):
		for player in range(len(playerIDs)):
			sheet.cell(row=player + 2, column=stat + 13, value=moneyStats[stat][player]) # averaged positions

	for player in range(len(playerIDs)):
		sheet.cell(row=player + 2, column=15, value=handsPlayed[player]) # averaged positions

	row = player + 2 + 1

	while row < 15: # Hopefully there won't be more than 25 players in one session!
		sheet.delete_rows(row) # delete rows that may remain from previous sessions
		row += 1


	wb.save(wb_path)




