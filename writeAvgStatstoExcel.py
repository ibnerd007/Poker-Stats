import openpyxl
from search import *
from average import *
import time

def writeAvgStatstoExcel(vpipM, pfrM, tbpM, cbpCountM, afM, afqM, wtsdM, wasdM, wasdRelM, mwas, mwbs, 
  		 			 ledgerM, playerIndices, handsPlayed, date, handTypeDesired):
	# Much like bankrolls, creates a weighted average of stats through total hands played
	# Calculates separately for Holdem and PLO, as well as combined

	# Read:    playerName = sheet.cell(row=2, column=1).value
	# Write:   sheet.cell(row=i + 2, column=1, value=playerDict[playerIDs[i]])

	# 1. Access workbook and correct sheet ----------------------------------------------------------

	wb_path = r'Outputs\stat averages.xlsx'

	wb = openpyxl.load_workbook(wb_path) # load existing workbook

	if handTypeDesired == 'NL':    sheet = wb['NL Stats-all sessions' ] # access sheet
	elif handTypeDesired == 'PLO': sheet = wb['PLO Stats-all sessions']
	else:                          sheet = wb['All Stats-all sessions']

	# 2. Establish players of interest, and make sure date is not entered previously ----------------

	dates = []

	for i in range(5, sheet.max_row + 1):
		cellDate = sheet.cell(row=i, column=1).value
		dates.append(cellDate)
	# dates is now filled

	if search(dates, date) != -1: # data from this date has been entered previously
		return
	else: # add date to column
		print('Adding average stat data...\n')
		if dates == [None]:
			sheet.cell(row=sheet.max_row, column=1, value=date)
		else:
			sheet.cell(row=sheet.max_row+1, column=1, value=date)

	# 4. Fill sheet with data ----------------------------------------------------------------------

	tdStats = (vpipM, pfrM, tbpM, afqM, wtsdM, wasdM, wasdRelM)

	# Fill Excel spreadsheet with percent stat data
	for stat in range(len(tdStats)): # cols
		for row, pI in enumerate(playerIndices): # rows
			# Assign player's index this session

			if pI != -1:
				# Get stats from this session and averaged sessions
				statAvg = sheet.cell(row=row+2, column=stat+3).value
				statThisSession  = tdStats[stat][pI][2]

				# Get hands played this session and total across all sessions
				totalHandsPlayed = sheet.cell(row=row+2, column=14).value
				handsPlayedThisSession  = handsPlayed[0][pI] + handsPlayed[1][pI]

				# Calculate weighted average
				newStatAvg = average(statAvg, statThisSession, totalHandsPlayed, handsPlayedThisSession)
				
				sheet.cell(row=row+2, column=stat+3, value=newStatAvg) # fill percent data
				sheet.cell(row=row+2, column=stat+3).number_format = '0.0%'


	# Fill Excel spreadsheet with C-bets vs opportunities and aggression factors, not a percent-based stats
	for row, pI in enumerate(playerIndices): # rows

		if pI != -1:
		
			totalHandsPlayed = sheet.cell(row=row+2, column=14).value
			handsPlayedThisSession  = handsPlayed[0][pI] + handsPlayed[1][pI]

			# ----------------------------------------------------------------------------------
			afCurr = afM[pI][2]
			
			if afCurr != -1: # if AF is not -1 (the player called at least once during session)
				afPrev = sheet.cell(row=row+2, column=11).value
				afAvg = average(afPrev, afCurr, totalHandsPlayed, handsPlayedThisSession)

				sheet.cell(row=row+2, column=11, value=afAvg)

			# -----------------------------------------------------------------------------------
			
			cbpBetPrev = sheet.cell(row=row+2, column=12).value
			cbpBetCurr = cbpCountM[pI][0] + cbpCountM[pI][1]

			cbpBetTotal = cbpBetPrev + cbpBetCurr

			sheet.cell(row=row+2, column=12, value=cbpBetTotal) # fill c-bets

			# -----------------------------------------------------------------------------------

			cbpOppsPrev = sheet.cell(row=row+2, column=13).value
			cbpOppsCurr = cbpCountM[pI][2] + cbpCountM[pI][3]

			cbpOppsTotal = cbpOppsPrev + cbpOppsCurr

			sheet.cell(row=row+2, column=13, value=cbpOppsTotal) # fill c-bet opportunities

			# -----------------------------------------------------------------------------------

			try:    cbp = cbpBetTotal/cbpOppsTotal
			except: cbp = 0

			sheet.cell(row=row+2, column=10, value=cbp) # fill percent data
			sheet.cell(row=row+2, column=10).number_format = '0.0%'

			# -----------------------------------------------------------------------------------

	for row, pI in enumerate(playerIndices):

		if pI != -1:
			prevTotal = sheet.cell(row=row+2, column=14).value
			currTotal = handsPlayed[0][pI] + handsPlayed[1][pI]

			sheet.cell(row=row+2, column=14, value=prevTotal + currTotal)

	wb.save(wb_path)
