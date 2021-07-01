import openpyxl
from search import *
import time

def writeBankrollsToExcel(ledgerM, playerIDs, playerNames, dateFormat, playerIndices):
	# Keeps a running bankroll of regular players across multiple sessions
	# Players tracked:
	# Fish, Raymond, Scott, Cedric

	# Read:    playerName = sheet.cell(row=2, column=1).value
	# Write:   sheet.cell(row=i + 2, column=1, value=playerDict[playerIDs[i]])

	""" PLEASE READ: sheet.max_row does not function as expected. Even if there is no value
	in an entire row, those cells can still trigger the max_row object. To reduce max_row, you
	must right click and delete the entire row(s), NOT just the content by pressing 'Delete'.
	Charts and graphs should move up correspondingly; this is how you know you have successfully
	lowered the max_row object. This can cause plenty of errors so be careful when you use it. """

	# 1. Open workbook and sheet -------------------------------------------------------------

	wb_path = r'Outputs\bankrolls.xlsx'

	wb = openpyxl.load_workbook(wb_path) # load existing workbook

	sheet = wb['Bankrolls'] # access sheet

	# 2. Check to make sure this session has not already been recorded -----------------------
	#    Compare this session's date with date column

	dates = []
	netCols = (3, 8, 13, 18) # columns in Excel where the net will be stored

	for i in range(2, sheet.max_row + 1):
		cellDate = sheet.cell(row=i, column=1).value
		dates.append(cellDate)

	if search(dates, dateFormat) != -1: # data from this date has been entered previously
		return

	else: # add date to column
		print('Adding bankroll data...\n')
		if dates == [None]:
			sheet.cell(row=sheet.max_row, column=1, value=dateFormat)
		else:
			sheet.cell(row=sheet.max_row+1, column=1, value=dateFormat)

	# 3. Get each player's index ------------- ------------------------------------------------

	for index in playerIndices:

	# 4. Add net to current net for each player -----------------------------------------------

		prevNet = sheet.cell(row=sheet.max_row-1, column=netCols[i]).value

		assert isinstance(prevNet, float) or isinstance(prevNet, int), \
		'prevNet is of {} (value = "{}"), incorrectly'.format(type(prevNet), prevNet)

		if index != -1: # player played the session. Else, net remains the same as previously
			newNet = ledgerM[index][2] + prevNet

			bankroll = sheet.cell(row=sheet.max_row-1, column=netCols[i] + 1).value
			ownMoneyInvested = sheet.cell(row=sheet.max_row-1, column=netCols[i] + 2).value
			buyIn = ledgerM[index][0]

			if buyIn > bankroll:
				ownMoneyInvested += (buyIn - bankroll)
				bankroll += (buyIn - bankroll) # update bankroll if more money was added to play

			net = ledgerM[index][2] # just set above

			bankroll += net

			assert bankroll >= 0, 'Bankroll = {.2f} < 0 for index {}'.format(bankroll, index)

			sheet.cell(row=sheet.max_row, column=netCols[i], value=newNet)
			sheet.cell(row=sheet.max_row, column=netCols[i] + 1, value=bankroll) # set bankroll
			sheet.cell(row=sheet.max_row, column=netCols[i] + 2, value=ownMoneyInvested) # set own money invested


		else: # show that player didn't play and fill columns

			UNCHnet              = sheet.cell(row=sheet.max_row-1, column=netCols[i]).value
			UNCHbankroll         = sheet.cell(row=sheet.max_row-1, column=netCols[i] + 1).value
			UNCHownMoneyInvested = sheet.cell(row=sheet.max_row-1, column=netCols[i] + 2).value

			sheet.cell(row=sheet.max_row, column=netCols[i] - 1, value='Didn\'t play')
			sheet.cell(row=sheet.max_row, column=netCols[i], value=UNCHnet)
			sheet.cell(row=sheet.max_row, column=netCols[i] + 1, value=UNCHbankroll) # set bankroll
			sheet.cell(row=sheet.max_row, column=netCols[i] + 2, value=UNCHownMoneyInvested) # all values are unchanged from previous


	wb.save(wb_path)